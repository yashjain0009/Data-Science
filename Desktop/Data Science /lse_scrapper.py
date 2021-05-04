from selenium import webdriver
from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
import openpyxl
import numpy as np
import requests
from bs4 import BeautifulSoup
import time
from bs4 import BeautifulSoup
path='/Users/yashjain/Downloads/chromedriver'
driver=webdriver.Chrome(path)
faculty_list=driver.get("https://www.lse.ac.uk/economics/people/faculty")
url="https://www.lse.ac.uk/economics/people/faculty"
source=requests.get(url).text
faculty_soup=BeautifulSoup(source,'lxml')
links=faculty_soup.findAll('div',class_='accordion__txt')

y=0
faculty_name_list=[]
website=[]
email=[]
cv=[]
position=[]
dept=[]
name=[]
expertise=[]

wb=openpyxl.Workbook()
ws1 = wb.create_sheet("London School of Economics ")
ws1.title = "LSE"

for link in links:
    faculty_list = link.find('a')
    faculty_name_list.append(faculty_list.text)

#search=driver.find_element_by_class_name('siteSearchTrigger')
#search.send_keys("masters")
#search.send_keys(return keys)

#link=soup.findAll('div',class_='accordion__txt')
for i in range (0,len(faculty_name_list)):
    #links = driver.find_elements_by_class_name("accordion__txt")
    #links2=np.array
    ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
    #print link.text

    #print y
    #WebDriverWait(driver,5,ignored_exceptions=ignored_exceptions).until(
    #    EC.presence_of_element_located((By.CLASS_NAME,"accordion__img"))
    #)
    #x = link.find_element_by_tag_name('a').text
    y+=1

    try:

        mai = WebDriverWait(driver, 5,ignored_exceptions=ignored_exceptions).until(
            EC.presence_of_element_located((By.LINK_TEXT, faculty_name_list[i]))

        )
        mai.click()
        urlpage = driver.current_url
        #print urlpage
        source = requests.get(urlpage).text
        soup = BeautifulSoup(source, 'lxml')
        basic_info = soup.find('div', class_='column__3 mediumColumn__6 largeColumn__8')
        contact_details = soup.findAll('a', class_="peopleContact__link")
        name.append(basic_info.h1.text)
        position.append(basic_info.h2.text)
        dept.append(basic_info.h3.text)
        for contact in contact_details:
            # print contact.attrs
            if contact.text == "Website":
                website.append(contact.attrs['href'])
            if contact.attrs['href'].startswith("mailto:"):
                email.append(contact.attrs['href'][7:])
        desc = soup.find('div', class_='people__bio')
        cv.append((desc.find('a', href=True)).attrs['href'])
        expertise.append((soup.find('div', class_='expertise')).text)


        driver.back()
        y=y+1
    except:
        driver.back()
        pass
counter=0
for i in range(1,len(faculty_name_list)-1):
    counter+=1
    try:
        c1=ws1.cell(row=i,column=1)
        c1.value=name[i]
    except:
        print (i)

    try:
        c2 = ws1.cell(row=i, column=2)
        c2.value = position[i]
    except:
        print (i)
        print (name[i])
    try:
        c3 = ws1.cell(row=i, column=3)
        c3.value = dept[i]
    except:
        print (i)
        print (name[i])
    try:
        c4=ws1.cell(row=i,column=4)
        c4.value = cv[i]
    except:
        print (name[i])
    try:
        c5 = ws1.cell(row=i, column=5)
        c5.value = website[i]
    except:
        print (i)
        print (name[i])
    try:
        c6 = ws1.cell(row=i, column=6)
        c6.value = email[i]
    except:
        print (i)
        print (name[i])
    try:
        c7 = ws1.cell(row=i, column=7)
        c7.value = cv[i]
    except:
        print (i)
        print (name[i])
    try:
        c8 = ws1.cell(row=i, column=7)
        c8.value = expertise[i]
    except:
        print (name[i])
    print (counter)
wb.template=False
wb.save(filename="Forgien.xlsx")
print (name)
print (position)
print (dept)
print (website)
print (email)
print (cv)
print (expertise)







#soup=BeautifulSoup(x,'lmxl')
#x=soup.find('div',class_='siteSearch__container')
#print x
#print (driver.title)
#time.sleep(5)
#driver.quit()
